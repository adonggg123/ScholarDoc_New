import openpyxl
import os

file_path = r"c:\Users\judej\ScholarDoc\0_Annex 5 - TES New Billing Form (1).xlsx"
print(f"File exists: {os.path.exists(file_path)}")

if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"Sheet '{name}' dimensions: {sheet.dimensions}")
        # print first few rows/cols to see the header structure
        for row in list(sheet.iter_rows(values_only=True))[:20]:
            if any(row):
                # Filter out completely empty rows for display
                print([str(cell)[:30] if cell is not None else "" for cell in row[:15]])
