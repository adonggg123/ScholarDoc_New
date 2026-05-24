import openpyxl

file_path = r"c:\Users\judej\ScholarDoc\0_Annex 5 - TES New Billing Form (1).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

for name in ['Annex 5-TES New Form 1', 'Annex 5-TES New Form 2', 'Annex 5-TES New Form 3']:
    sheet = wb[name]
    print(f"\n===== Non-empty cells in {name} =====")
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                # If it's a string and starts with '=', or if it's in rows 40-70
                if isinstance(val, str) and val.startswith('='):
                    print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
                elif r >= 40:
                    print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
